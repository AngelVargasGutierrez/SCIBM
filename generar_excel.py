import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Por favor instala openpyxl ejecutando: pip install openpyxl")
    exit(1)

def parse_markdown_table(text):
    tables = []
    lines = text.strip().split('\n')
    current_table = []
    in_table = False
    
    for line in lines:
        if line.strip().startswith('|'):
            in_table = True
            current_table.append(line.strip())
        else:
            if in_table:
                tables.append(current_table)
                current_table = []
                in_table = False
    if current_table:
        tables.append(current_table)
        
    parsed_tables = []
    for t in tables:
        rows = []
        for row in t:
            if re.match(r'^\|[\s\-\|]+\|$', row):
                continue
            cols = [col.strip() for col in row.split('|')[1:-1]]
            rows.append(cols)
        parsed_tables.append(rows)
    return parsed_tables

def process_files(directory):
    print(f"Leyendo archivos .md de la carpeta: {directory}...")
    wb = Workbook()
    wb.remove(wb.active)
    
    fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    
    font_bold = Font(bold=True, color="000000")
    font_white_bold = Font(bold=True, color="FFFFFF")
    font_normal = Font(color="000000")
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    archivos_procesados = 0

    for filename in sorted(os.listdir(directory)):
        if filename.endswith(".md"):
            filepath = os.path.join(directory, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            
            tables = parse_markdown_table(content)
            if not tables:
                continue
            
            sheet_title = filename.replace('.md', '')
            ws = wb.create_sheet(title=sheet_title)
            
            # Anchos de columna basados en la plantilla visual
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 50
            
            row_idx = 1
            
            # Buscar el título real en el markdown (ej. "a) UC-001 - Iniciar Sesión")
            # Podemos inferirlo del primer "Caso de uso" en la tabla 1
            caso_de_uso_text = sheet_title
            for row in tables[0]:
                if len(row) >= 2 and "Caso de uso" in row[0]:
                    caso_de_uso_text = row[1].replace('**', '')
                    break
            
            # 1. Título superior
            c_title = ws.cell(row=row_idx, column=1, value=f"a) {caso_de_uso_text}")
            c_title.font = font_bold
            row_idx += 2
            
            # 2. Detalles (Tabla 1)
            for row in tables[0]:
                if len(row) >= 2:
                    c1 = ws.cell(row=row_idx, column=1, value=row[0].replace('**', ''))
                    c1.font = font_normal
                    c1.fill = fill_gray
                    c1.border = border_thin
                    c1.alignment = Alignment(vertical="top")
                    
                    c2 = ws.cell(row=row_idx, column=2, value=row[1])
                    c2.font = font_normal
                    c2.border = border_thin
                    c2.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    row_idx += 1
            
            # Buscar los títulos de los flujos alternos
            lines = content.split('\n')
            titulos_flujos = []
            for line in lines:
                if line.startswith('### '):
                    titulos_flujos.append(line.replace('### ', '').strip())
            
            # 3. Cursos de Eventos (Tabla 2 en adelante)
            for idx, table in enumerate(tables[1:]):
                title = titulos_flujos[idx] if idx < len(titulos_flujos) else f"Flujo alterno {idx}"
                
                # Fila de Cabecera Negra fusionada
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                c_header = ws.cell(row=row_idx, column=1, value=title)
                c_header.fill = fill_black
                c_header.font = font_white_bold
                c_header.alignment = Alignment(horizontal="center", vertical="center")
                # Aplicar bordes a las celdas combinadas
                ws.cell(row=row_idx, column=1).border = border_thin
                ws.cell(row=row_idx, column=2).border = border_thin
                row_idx += 1
                
                # Fila de Subcabecera Gris ("Usuario" | "Sistema")
                c_u = ws.cell(row=row_idx, column=1, value="Usuario")
                c_u.fill = fill_gray
                c_u.font = font_bold
                c_u.border = border_thin
                c_u.alignment = Alignment(vertical="center")
                
                c_s = ws.cell(row=row_idx, column=2, value="Sistema")
                c_s.fill = fill_gray
                c_s.font = font_bold
                c_s.border = border_thin
                c_s.alignment = Alignment(vertical="center")
                row_idx += 1
                
                # Procesar pasos y agrupar Usuario y Sistema en la misma fila si es posible
                steps = table[1:] # Saltar cabecera de la tabla markdown
                
                i = 0
                while i < len(steps):
                    step_data = steps[i]
                    # La estructura es [# , Usuario, Sistema]
                    if len(step_data) >= 3:
                        num = step_data[0].strip()
                        user_txt = step_data[1].strip()
                        sys_txt = step_data[2].strip()
                        
                        # Si es una acción de usuario y no de sistema
                        if user_txt != "" and sys_txt == "":
                            # Mirar el siguiente paso para ver si es del sistema
                            if i + 1 < len(steps):
                                next_step = steps[i+1]
                                if len(next_step) >= 3:
                                    next_num = next_step[0].strip()
                                    next_user = next_step[1].strip()
                                    next_sys = next_step[2].strip()
                                    
                                    if next_user == "" and next_sys != "":
                                        # ¡Agrupar en la misma fila!
                                        val_u = f"{num}. {user_txt}"
                                        val_s = f"{next_num}. {next_sys}"
                                        
                                        c1 = ws.cell(row=row_idx, column=1, value=val_u)
                                        c1.border = border_thin
                                        c1.alignment = Alignment(wrap_text=True, vertical="top")
                                        
                                        c2 = ws.cell(row=row_idx, column=2, value=val_s)
                                        c2.border = border_thin
                                        c2.alignment = Alignment(wrap_text=True, vertical="top")
                                        
                                        row_idx += 1
                                        i += 2 # Saltar el siguiente paso
                                        continue
                        
                        # Si no se agruparon, escribir en su propia fila
                        val_u = f"{num}. {user_txt}" if user_txt else ""
                        val_s = f"{num}. {sys_txt}" if sys_txt else ""
                        
                        c1 = ws.cell(row=row_idx, column=1, value=val_u)
                        c1.border = border_thin
                        c1.alignment = Alignment(wrap_text=True, vertical="top")
                        
                        c2 = ws.cell(row=row_idx, column=2, value=val_s)
                        c2.border = border_thin
                        c2.alignment = Alignment(wrap_text=True, vertical="top")
                        
                        row_idx += 1
                    i += 1
            
            archivos_procesados += 1
                
    if archivos_procesados > 0:
        output_file = "Requerimientos_Narrativa.xlsx"
        wb.save(output_file)
        print(f"\n¡Éxito! Se han procesado {archivos_procesados} archivos .md.")
        print(f"El archivo Excel ha sido generado aplicando la PLANTILLA OFICIAL: {output_file}")
    else:
        print("No se encontraron tablas en los archivos .md.")

if __name__ == "__main__":
    import sys
    base_dir = os.path.dirname(os.path.abspath(__file__))
    directory = os.path.join(base_dir, "NARRATIVA")
    
    if not os.path.exists(directory):
        print(f"Error: La carpeta {directory} no existe.")
    else:
        process_files(directory)
